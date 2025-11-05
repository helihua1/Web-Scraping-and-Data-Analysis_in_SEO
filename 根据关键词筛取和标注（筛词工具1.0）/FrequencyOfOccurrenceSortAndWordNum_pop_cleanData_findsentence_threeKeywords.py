import jieba
from collections import defaultdict
import pandas as pd
import numpy as np
from tqdm import tqdm
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import openpyxl

"""
seo工具：在一组数据中，聚合相似的句子


目的：excel一列中文句子中，聚合相似的句子，找到共同的关键词/关键句
处理顺序：
1，jieba分词，
2，然后n-gram拼成 1-n 个词组合的字符串，
3，去掉需要忽略的词，
4，根据词的长度（len_rate）和在句子中出现的频率(num_rate)作为权重，计算全部数据中每个词的权重，
5，找到每条数据中权重最高的1个词，作为这条数据中的关键词

输入内容如下：
#长度倍率在此处输入：
    len_rate = 1.7 
#词频倍率在此处输入：
    num_rate_Input = 1.5 
#input_excel需要分析的文档 ：
    input_excel = r"D:\sort\10.31分析seo策略\数据分析.xlsx"
# output_excel 是输出文档：
    output_excel = fr'D:\sort\10.31分析seo策略\数据分析_3词_长度倍率{len_rate}X_词频倍率{num_rate_Input}X.xlsx'
# source_path是需要不让其参与聚合的词：
    source_path = r"D:\sort\10.31分析seo策略\数据清洗词.xlsx"  
"""



def get_words_to_remove(source_path):
    """从源Excel获取第一列的需要忽略的词列表"""
    df = pd.read_excel(source_path, header=None)
    words = df.iloc[:, 0].dropna().astype(str).unique().tolist()
    print("需要删除的词列表：")
    for i, word in enumerate(words, 1):
        print(f"{i}. {word}")
    print(f"\n共找到 {len(words)} 个需要删除的词")
    return words

def preprocess_text(text):
    """预处理文本：分词并保留2字及以上词语"""
    if not isinstance(text, str):
        text = str(text)
    return [word for word in jieba.lcut(text) if len(word) >= 2]


def calculate_advanced_weights(sentences, rate, num_rate, words_to_remove=None, ngram_range=(1, 3)):
    """
    改进版版权重计算：
    - 支持 n-gram（1~n词组合）
    - 考虑词长加权
    - 输出每个句子的前3高权重词及其权重
    """
    if words_to_remove is None:
        words_to_remove = []

    word_stats = defaultdict(lambda: {'count': 0, 'length': 0})

    # 第一次遍历：统计所有 n-gram
    for sentence in tqdm(sentences, desc="构建词库"):
        words = preprocess_text(sentence)
        n = len(words)
        for k in range(ngram_range[0], ngram_range[1] + 1):
            for i in range(n - k + 1):
                ngram = ''.join(words[i:i + k])
                if len(ngram) >= 2:
                    word_stats[ngram]['count'] += num_rate
                    word_stats[ngram]['length'] = len(ngram)

    # 移除包含停用词的 key
    for w in words_to_remove:
        keys_to_remove = [key for key in word_stats if w in key]
        for key in keys_to_remove:
            word_stats.pop(key, None)

    # 计算加权频率
    weighted_word_freq = {}
    for word, stats in word_stats.items():
        weight = stats['count'] * (rate ** (stats['length'] - 2))
        weighted_word_freq[word] = weight

    # 第二次遍历：每个句子找前3高权重词
    sentence_records = []
    for sentence in tqdm(sentences, desc="计算权重"):
        words = preprocess_text(sentence)
        if not words:
            sentence_records.append({
                'sentence': sentence,
                'keyword1': '', 'weight1': 0,
                'keyword2': '', 'weight2': 0,
                'keyword3': '', 'weight3': 0,
            })
            continue

        ngrams = []
        for k in range(ngram_range[0], ngram_range[1] + 1):
            for i in range(len(words) - k + 1):
                ngram = ''.join(words[i:i + k])
                if len(ngram) >= 2:
                    ngrams.append(ngram)

        # 找出前3高权重词
        weights = [(ngram, weighted_word_freq.get(ngram, 0)) for ngram in ngrams]
        weights = sorted(weights, key=lambda x: x[1], reverse=True)[:3]

        # 补足3个
        while len(weights) < 3:
            weights.append(('', 0))

        sentence_records.append({
            'sentence': sentence,
            'keyword1': weights[0][0], 'weight1': weights[0][1],
            'keyword2': weights[1][0], 'weight2': weights[1][1],
            'keyword3': weights[2][0], 'weight3': weights[2][1],
        })

    return pd.DataFrame(sentence_records), weighted_word_freq


def sort_sentences(df):
    """多级排序：支持三关键词权重+长度+句子排序"""
    # 先生成临时列保存关键词长度
    for i in range(1, 4):
        df[f'keyword{i}_len'] = df[f'keyword{i}'].str.len()

    # 排序规则：weight1 > keyword1_len > weight2 > keyword2_len > weight3 > keyword3_len > sentence
    sort_cols = [
        'weight1', 'keyword1_len',
        'weight2', 'keyword2_len',
        'weight3', 'keyword3_len',
        'sentence'
    ]
    ascending_flags = [False, False, False, False, False, False, True]

    sorted_df = df.sort_values(by=sort_cols, ascending=ascending_flags)

    # 删除临时列
    df.drop(columns=[f'keyword{i}_len' for i in range(1, 4)], inplace=True)
    return sorted_df


def main(input_path, output_path, rate, num_rate, words_to_remove):
    """主处理函数，含Excel列宽自动调整与首列背景色"""
    df_input = pd.read_excel(input_path, header=None)
    sentences = df_input.iloc[:, 0].astype(str).tolist()

    result_df, _ = calculate_advanced_weights(sentences, rate, num_rate, words_to_remove)
    sorted_df = sort_sentences(result_df)

    # 输出到Excel
    sorted_df[['sentence',
               'keyword1', 'weight1',
               'keyword2', 'weight2',
               'keyword3', 'weight3']].to_excel(
        output_path, index=False,
        header=['句子', '关键词1', '权重1', '关键词2', '权重2', '关键词3', '权重3']
    )

    # 使用 openpyxl 打开Excel调整列宽和首列背景色
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # 设置首列背景色
    fill_color = PatternFill(start_color='F4F9F5', end_color='F4F9F5', fill_type='solid')
    for row in ws.iter_rows(min_row=1, max_col=1, max_row=ws.max_row):
        for cell in row:
            cell.fill = fill_color

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        # 适当加一点空间
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_path)
    print(f"处理完成，结果已保存到 {output_path}，列宽已自适应，首列背景已设置。")

if __name__ == "__main__":
    jieba.initialize()


    # # 过去
    # len_rate_Input = 1 #长度倍率在此处输入
    # len_rate = len_rate_Input * 1.5
    # print('长度倍率为' + str(len_rate_Input))


    #直接输入，大于1的数
    len_rate = 1.7 #长度倍率在此处输入
    print('长度倍率为' + str(len_rate))

    num_rate_Input = 1.5 #词频倍率在此处输入
    print('词频倍率为' + str(num_rate_Input))

    num_rate = num_rate_Input

    input_excel = r"D:\sort\10.31分析seo策略\数据分析.xlsx"
    output_excel = fr'D:\sort\10.31分析seo策略\数据分析_3词_长度倍率{len_rate}X_词频倍率{num_rate_Input}X.xlsx'
    source_path = r"D:\sort\10.31分析seo策略\数据清洗词.xlsx"  # 包含需要移除的词
    words_to_remove = get_words_to_remove(source_path)



    main(input_excel, output_excel,len_rate,num_rate,words_to_remove)