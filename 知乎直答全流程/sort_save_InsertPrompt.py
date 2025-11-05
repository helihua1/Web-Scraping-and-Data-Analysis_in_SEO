import re

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill



def process_excel_files():
    # 文件路径设置
    sorted_file_path = r"D:\sort\selenium文档\10w中 白癜风待插入提示词-百科只要带白癜风词汇的.xlsx"
    priority_file_path = r"D:\sort\selenium文档\zbw 生成提示词-白癜风.xlsx"
    output_file_path = r"D:\sort\selenium文档\提示词插入结果_sort_save_InsertPrompt.xlsx"

    try:
        #
        sorted_df = pd.read_excel(sorted_file_path, header=None, usecols=[0, 1, 2, 3, 4])
        
        # 获取前三列数据
        # scores = sorted_df[0].tolist()  # 第0列: score
        type1 = sorted_df[0].tolist()  
        type2 = sorted_df[1].tolist()  
        sentences = sorted_df[2].tolist()  # 第3列: 需要匹配的sentences
        type3 = sorted_df[3].tolist()  
        user_inputs = sorted_df[4].tolist()  # 第4列: user_input

        #从第二列开始的两列
        # priority_df = pd.read_excel(priority_file_path, header=None, usecols=[0, 1, 2, 3, 4, 5, 6, 7])
        priority_df = pd.read_excel(priority_file_path, header=None, usecols=[1, 2])

        # 存储第4列关键词及其属性（前3列）
        priority1_words = []
        priority1_attrs = {}  # 键:关键词, 值:属性列表[attr0, attr1, attr2]
        # # 存储第8列关键词及其属性（前3列）
        # priority2_words = []
        # priority2_attrs = {}  # 键:关键词, 值:属性列表[attr4, attr5, attr6]

        # 提取关键词及其对应的属性

        for idx, row in priority_df.iterrows():

            # 处理第二列的提示词及其属性
            word1 = row[1]
            if not pd.isna(word1):
                priority1_words.append(word1)
                # 获取提示词
                priority1_attrs[word1] = [row[2]]

            # # 处理第8列（索引7）的关键词及其属性
            # word2 = row[7]
            # if not pd.isna(word2):
            #     priority2_words.append(word2)
            #     # 获取前3列属性（索引4、5、6）
            #     priority2_attrs[word2] = [row[4], row[5], row[6]]

        # 用于存储匹配到的句子及对应的关键词、属性和原始数据
        matched_dict = {}  # 键:关键词, 值:列表，每个元素为[句子, score, user_input]
        remaining_user_inputs = []  # 存储未匹配的user_input

        # 遍历所有句子，检查是否包含优先聚合词并记录匹配的关键词
        for i, sentence in enumerate(sentences):
            if pd.isna(sentence):  # 不对空值进行比对
                # 即使句子为空，也保留user_input
                if not pd.isna(user_inputs[i]):
                    remaining_user_inputs.append([type1[i], type2[i], sentence, type3[i], user_inputs[i]])
                continue

            matched_word = None
            # 先检查第4列优先级高的词
            for word in priority1_words:
                if str(word) == str(sentence):
                    matched_word = word
                    break

            # # 如果没有匹配到第4列的词，检查第8列的词
            # if matched_word is None:
            #     for word in priority2_words:
            #         if str(word) in str(sentence):
            #             matched_word = word
            #             break

            # 处理匹配结果
            if matched_word is not None:
                # 将句子及其对应的score和user_input添加到对应关键词的列表中
                if matched_word not in matched_dict:
                    matched_dict[matched_word] = []
                matched_dict[matched_word].append([type1[i], type2[i], sentence, type3[i], user_inputs[i]])
            else:
                # 未匹配的句子，只保留user_input
                if not pd.isna(user_inputs[i]):
                    remaining_user_inputs.append([type1[i], type2[i], sentence, type3[i], user_inputs[i]])

        # 按关键词优先级顺序聚合句子及相关信息
        aggregated_data = []  # 每个元素: [attr1, attr2, attr3, 关键词, 句子, score, user_input]
        
        # 先处理第4列的关键词（按原顺序）
        for word in priority1_words:
            if word in matched_dict:
                # 获取该关键词对应的属性
                attrs = priority1_attrs[word]
                # 为每个匹配的句子添加 插入的提示词(attrs[0])，提示词属性(word),
                for match in matched_dict[word]:
                    aggregated_data.append([ match[0], match[1], match[2], match[3],match[4],attrs[0], word])
                del matched_dict[word]  # 避免重复处理

        # # 再处理第8列的关键词（按原顺序）
        # for word in priority2_words:
        #     if word in matched_dict:
        #         # 获取该关键词对应的属性
        #         attrs = priority2_attrs[word]
        #         # 为每个匹配的句子添加属性、关键词、句子、score和user_input
        #         for match in matched_dict[word]:
        #             aggregated_data.append([attrs[0], attrs[1], attrs[2], word, match[0], match[1], match[2]])
        #         del matched_dict[word]

        # 准备构建结果DataFrame的数据
        max_length = max(len(aggregated_data), len(remaining_user_inputs))

        # 拆分聚合数据到各列
        remaining_user_inputs1 = [item[0] for item in remaining_user_inputs] + [None] * (max_length - len(remaining_user_inputs))
        remaining_user_inputs2 = [item[1] for item in remaining_user_inputs] + [None] * (max_length - len(remaining_user_inputs))
        remaining_user_inputs3 = [item[2] for item in remaining_user_inputs] + [None] * (max_length - len(remaining_user_inputs))
        remaining_user_inputs4 = [item[3] for item in remaining_user_inputs] + [None] * (max_length - len(remaining_user_inputs))
        remaining_user_inputs5 = [item[4] for item in remaining_user_inputs] + [None] * (max_length - len(remaining_user_inputs))

        agg_type1 = [item[0] for item in aggregated_data] + [None] * (max_length - len(aggregated_data))
        agg_type2 = [item[1] for item in aggregated_data] + [None] * (max_length - len(aggregated_data))
        agg_sentences = [item[2] for item in aggregated_data] + [None] * (max_length - len(aggregated_data))
        agg_type3 = [item[3] for item in aggregated_data] + [None] * (max_length - len(aggregated_data))
        agg_user_inputs = [item[4] for item in aggregated_data] + [None] * (max_length - len(aggregated_data))
        agg_prompt = [item[5] for item in aggregated_data] + [None] * (max_length - len(aggregated_data))
        agg_prompt_type = [item[6] for item in aggregated_data] + [None] * (max_length - len(aggregated_data))

        # 创建新的DataFrame，列分布：
        # 0: 未匹配的user_input
        # 1: 预留列
        # 2: 预留列
        # 3: 关键词属性1
        # 4: 关键词属性2
        # 5: 关键词属性3
        # 6: 匹配到的关键词
        # 7: 匹配到的句子
        # 8: 匹配到的score
        # 9: 匹配到的user_input
        result_data = {
            0: remaining_user_inputs1,
            1: remaining_user_inputs2, 
            2: remaining_user_inputs3, 
            3: remaining_user_inputs4, 
            4: remaining_user_inputs5, 
            5: [None] * max_length,  
            6: agg_type1, 
            7: agg_type2, 
            8: agg_sentences, 
            9:  agg_type3,
            10: agg_user_inputs,  
            11: agg_prompt, 
            12: agg_prompt_type  
        }

        result_df = pd.DataFrame(result_data)

        # 保存结果到新的Excel文件
        result_df.to_excel(output_file_path, index=False, header=False)
        
        # 设置第5列（索引为5）的背景色为灰色
        wb = load_workbook(output_file_path)
        ws = wb.active
        
        # 创建灰色填充样式
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # 为第5列的所有单元格设置灰色背景
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=6).fill = gray_fill  # 第5列对应Excel的第6列（因为从1开始计数）
        
        # 保存修改后的文件
        wb.save(output_file_path)
        
        print(f"处理完成！结果已保存至：{output_file_path}")
        print("第5列已设置为灰色背景")

    except Exception as e:
        print(f"处理过程中发生错误：{str(e)}")


if __name__ == "__main__":
    process_excel_files()