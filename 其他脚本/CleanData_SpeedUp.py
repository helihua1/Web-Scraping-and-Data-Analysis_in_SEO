import pandas as pd
import re
from openpyxl import load_workbook
import concurrent.futures
from typing import List, Tuple
import threading
import os
'''
clean_source_path文档中第一列的词，在clean_target_path文档中删除。
优化版本：使用多线程并行处理，提高处理速度。
'''

def load_config(config_path='config.txt'):
    """从配置文件加载路径配置"""
    config = {}
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    config[key.strip()] = value.strip()
        return config
    except FileNotFoundError:
        print(f"配置文件 {config_path} 未找到，使用默认路径")
        return {}
    except Exception as e:
        print(f"读取配置文件时出错: {e}")
        return {}


def get_words_to_remove(source_path):
    """从源Excel获取第一列的去重词列表"""
    df = pd.read_excel(source_path, header=None)
    words = df.iloc[:, 0].dropna().astype(str).unique().tolist()
    print("需要删除的词列表：")
    for i, word in enumerate(words, 1):
        print(f"{i}. {word}")
    print(f"\n共找到 {len(words)} 个需要删除的词")
    return words


def clean_excel(target_path, output_path, words):
    """执行实际的删除操作（包括句子中的词）- 优化版本"""
    wb = load_workbook(target_path)
    
    # 预编译正则表达式，提高性能
    if words:
        # 将所有词组合成一个正则表达式，用|分隔
        pattern = '|'.join(re.escape(word) for word in words)
        regex = re.compile(pattern, flags=re.IGNORECASE)
    else:
        regex = None

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if regex:
                        # 一次性替换所有匹配的词
                        cell.value = regex.sub('', str(cell.value))
                        # 清理多余空格
                        cell.value = ' '.join(str(cell.value).split())

    wb.save(output_path)


def clean_excel_fast(target_path, output_path, words, max_workers=4):
    """执行实际的删除操作（包括句子中的词）- 多线程优化版本"""
    wb = load_workbook(target_path)
    
    # 预编译正则表达式
    if words:
        pattern = '|'.join(re.escape(word) for word in words)
        regex = re.compile(pattern, flags=re.IGNORECASE)
    else:
        regex = None
    
    def process_cell(cell_data):
        """处理单个单元格的函数"""
        sheet_name, row_idx, col_idx, value = cell_data
        if value and isinstance(value, str) and regex:
            # 一次性替换所有匹配的词
            cleaned_value = regex.sub('', str(value))
            # 清理多余空格
            cleaned_value = ' '.join(cleaned_value.split())
            return sheet_name, row_idx, col_idx, cleaned_value
        return sheet_name, row_idx, col_idx, value
    
    # 收集所有需要处理的单元格数据
    cells_to_process = []
    for sheet in wb.worksheets:
        for row_idx, row in enumerate(sheet.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str):
                    cells_to_process.append((sheet.title, row_idx, col_idx, cell.value))
    
    # 使用多线程处理
    processed_cells = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 分批处理，避免创建过多线程
        batch_size = 1000
        for i in range(0, len(cells_to_process), batch_size):
            batch = cells_to_process[i:i + batch_size]
            futures = [executor.submit(process_cell, cell_data) for cell_data in batch]
            for future in concurrent.futures.as_completed(futures):
                processed_cells.append(future.result())
    
    # 将处理结果应用到工作簿
    for sheet_name, row_idx, col_idx, cleaned_value in processed_cells:
        sheet = wb[sheet_name]
        sheet.cell(row=row_idx, column=col_idx, value=cleaned_value)
    
    wb.save(output_path)
    return len(processed_cells)


def main():
    # 从配置文件加载路径
    config = load_config()
    
    # 文件路径配置 - 优先使用配置文件中的路径，如果没有则使用默认路径
    source_path = config.get('clean_source_path', r'D:\sort\数据清洗需要剔去的常用语.xlsx')
    target_path = config.get('clean_target_path', r'D:\sort\A.xlsx')
    output_path = config.get('clean_output_path', r'D:\sort\A_cleaned.xlsx')

    # 检查文件是否存在
    if not os.path.exists(source_path):
        print(f"错误：源文件不存在: {source_path}")
        print("请检查配置文件中的 clean_source_path 设置")
        return
    
    if not os.path.exists(target_path):
        print(f"错误：目标文件不存在: {target_path}")
        print("请检查配置文件中的 clean_target_path 设置")
        return

    try:
        print("开始处理...")
        print(f"源文件: {source_path}")
        print(f"目标文件: {target_path}")
        print(f"输出文件: {output_path}")
        
        words_to_remove = get_words_to_remove(source_path)
        
        # 选择处理方式
        print("\n请选择处理方式：")
        print("1. 普通版本（适合少量词）")
        print("2. 快速版本（适合大量词，推荐）")
        
        choice = input("请输入选择 (1 或 2，默认选择 2): ").strip()
        
        if choice == "1":
            print("使用普通版本处理...")
            clean_excel(target_path, output_path, words_to_remove)
        else:
            print("使用快速版本处理...")
            # 根据CPU核心数自动调整线程数
            max_workers = min(os.cpu_count() or 4, 8)  # 最多8个线程
            print(f"使用 {max_workers} 个线程并行处理...")
            processed_count = clean_excel_fast(target_path, output_path, words_to_remove, max_workers)
            print(f"共处理了 {processed_count} 个单元格")
        
        print(f"\n处理完成！结果已保存到: {output_path}")

    except Exception as e:
        print(f"\n处理过程中出错: {str(e)}")


if __name__ == "__main__":
    main()
    input("按回车键退出...")