import pandas as pd
import re
from openpyxl import load_workbook
import os

"""
clean_source_path文档中第一列的词，在clean_target_path文档中删除。
"""
def load_config(config_path='config.txt'):
    """从配置文件加载路径配置"""
    config = {}
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    config[key.strip()] = value.strip()
        return config
    except FileNotFoundError:
        print(f"配置文件 {config_path} 未找到，使用默认配置")
        return {
            'source_path': r'D:\sort\数据清洗需要剔去的常用语.xlsx',
            'target_path': r'D:\sort\A.xlsx',
            'output_path': r'D:\sort\A_cleaned.xlsx'
        }


def get_words_to_remove(source_path):
    """从源Excel获取第一列的去重词列表"""
    df = pd.read_excel(source_path, header=None)
    words = df.iloc[:, 0].dropna().astype(str).unique().tolist()
    print("需要删除的词列表：")
    for i, word in enumerate(words, 1):
        print(f"{i}. {word}")
    print(f"\n共找到 {len(words)} 个需要删除的词")
    return words


def clean_excel(target_path, output_path, words):
    """执行实际的删除操作（包括句子中的词）"""
    wb = load_workbook(target_path)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for word in words:
                        # 直接删除所有匹配的词（包括句子中的词）
                        cell.value = re.sub(re.escape(word), '', str(cell.value), flags=re.IGNORECASE)
                    # 清理多余空格
                    cell.value = ' '.join(str(cell.value).split())

    wb.save(output_path)


def main():
    # 从配置文件加载路径配置
    config = load_config()
    
    source_path = config.get('source_path')
    target_path = config.get('target_path')
    output_path = config.get('output_path')
    
    print("当前配置：")
    print(f"源文件路径: {source_path}")
    print(f"目标文件路径: {target_path}")
    print(f"输出文件路径: {output_path}")
    print()

    try:
        print("开始处理...")
        words_to_remove = get_words_to_remove(source_path)
        clean_excel(target_path, output_path, words_to_remove)
        print(f"\n处理完成！结果已保存到: {output_path}")

    except Exception as e:
        print(f"\n处理过程中出错: {str(e)}")


if __name__ == "__main__":

    main()